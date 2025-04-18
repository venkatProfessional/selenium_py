from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium import webdriver
import time
import openpyxl
import os

from selenium_helper.common_code import CommonCode


# file --> workbook-->sheets--> Rows-->cells

class WriteDataFromExcel:
    def __init__(self):
        self.helper = CommonCode()
        self.driver = self.helper.driver

        # File path
        file = r"C:\Users\JothiVenkatajalapath\Desktop\Working_with_py\Book1.xlsx"

        try:
            # Check if file is accessible for writing
            if not os.access(file, os.W_OK):
                print(f"WARNING: You don't have write permission for {file}")
                print("Make sure the file isn't open in Excel and you have proper permissions")
                return

            workbook = openpyxl.load_workbook(file)
            sheet = workbook["Sheet3"]

            # Loops
            for i in range(1, 6):
                for j in range(1, 4):
                    sheet.cell(i, j).value = "Welcome"

            # writing different datas
            # this  is the way we need to insert a data

            sheet.cell(5,1).value = "Good"
            sheet.cell(5,2).value ="Bad"
            sheet.cell(5,3).value ="ugly"

            try:
                workbook.save(file)
                print("Excel file updated successfully")
            except PermissionError:
                # Try saving with a different filename if original is locked
                backup_file = file.replace('.xlsx', '_backup.xlsx')
                print(f"Unable to save to original file. Trying backup file: {backup_file}")
                workbook.save(backup_file)
                print(f"Data saved to backup file: {backup_file}")

        except Exception as e:
            print(f"Error: {type(e).__name__} - {e}")
        finally:
            print("Excel data operation finished")

    def close(self):
        self.driver.quit()


if __name__ == "__main__":
    test_script = WriteDataFromExcel()
    test_script.close()