from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium import webdriver
import time
import openpyxl

from selenium_helper.common_code import CommonCode


# file --> workbook-->sheets--> Rows-->cells

class ReadDataFromExcel:
    def __init__(self):
        self.helper = CommonCode()
        self.driver = self.helper.driver

        # ✅ Make sure to specify the full path to the Excel file
        file = r"C:\Users\JothiVenkatajalapath\Desktop\Working_with_py\Book1.xlsx"

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook["Sheet1"]
            rows = sheet.max_row
            col = sheet.max_column
            print(f"Max Rows: {rows}")
            print(f"Max Columns: {col}")
        except FileNotFoundError:
            print(f"File not found: {file}")
        except Exception as e:
            print(f"Error reading Excel file: {e}")

    #         Loops

        for i in range(1,rows+1):
            for j in range(1,col+1):
                 print(sheet.cell(i,j).value)
            print("")


    def close(self):
        self.driver.quit()


if __name__ == "__main__":
    test_script = ReadDataFromExcel()
    test_script.close()
