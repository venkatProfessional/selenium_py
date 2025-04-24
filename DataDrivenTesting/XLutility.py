# handling excel files through pycharm is called data driven testing
# excelutils in excel utils everything is a function

import openpyxl
from openpyxl.styles import PatternFill

# Get the total number of rows in the sheet
def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row

# Get the total number of columns in the sheet
def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column

# Read a specific cell value
def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value

# Write data to a specific cell
def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)

# Optional: Write with cell highlighting
def fillGreenColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    sheet.cell(row=rownum, column=columnno).fill = fill
    workbook.save(file)

def fillRedColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    sheet.cell(row=rownum, column=columnno).fill = fill
    workbook.save(file)



